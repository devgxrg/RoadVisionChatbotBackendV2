
import re
import json
from typing import List, Dict, Tuple
import os
import traceback
import time
from pathlib import Path
from enum import Enum

# PDF Processing
try:
    from llama_parse import LlamaParse, ResultType
    import pytesseract
    from PIL import Image
    import fitz  # PyMuPDF
    import pdfplumber
    HAS_PDF_LIBS = True
except ImportError:
    HAS_PDF_LIBS = False

# Excel Processing
try:
    import pandas as pd
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    HAS_EXCEL_LIBS = True
except ImportError:
    HAS_EXCEL_LIBS = False

# HTML Processing
try:
    from bs4 import BeautifulSoup
    HAS_HTML_LIBS = True
except ImportError:
    HAS_HTML_LIBS = False

# Import from your app
from app.config import settings
from app.core.global_stores import upload_jobs
from app.modules.askai.models.document import ProcessingStage


# ============================================================================
# PDF PROCESSOR
# ============================================================================

class PDFProcessor:
    """Enhanced PDF processing with LlamaParse OCR"""
    
    job_id = None
    
    def __init__(self, embedding_model, tokenizer):
        self.embedding_model = embedding_model
        self.tokenizer = tokenizer
        
        if not HAS_PDF_LIBS:
            print("⚠️  PDF processing libraries not available")
            return
        
        llama_key = settings.LLAMA_CLOUD_API_KEY
        if llama_key:
            self.llama_parser = LlamaParse(
                api_key=llama_key,
                result_type=ResultType.MD,
                parsing_instruction="Extract all text, tables, and structure.",
                num_workers=2,
                verbose=False,
                max_timeout=600,
            )
            self.has_llamaparse = True
            print("✅ LlamaParse OCR initialized")
        else:
            self.has_llamaparse = False
            print("⚠️  LlamaParse not available")
    
    def clean_text(self, text: str) -> str:
        """Clean and normalize text"""
        if not text:
            return ""
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r'[^\w\s\.\,\;\:\!\?\-\(\)\[\]\"\'\/\@\#\$\%\&\*\+\=]', ' ', text)
        return text.strip()
    
    def _clean_metadata(self, metadata: Dict) -> Dict:
        """Clean metadata for ChromaDB compatibility"""
        cleaned = {}
        for k, v in metadata.items():
            str_val = str(v)
            str_val = re.sub(r'[^\w\s\-\.\,\/]', '_', str_val)
            str_val = str_val.strip()
            cleaned[k] = str_val if str_val else "unknown"
        return cleaned
    
    def update_progress(self, stage: ProcessingStage, progress: float) -> None:
        print(f"📄 Progress: {stage} {progress:.1f}%")
        if not self.job_id:
            return
        if self.job_id in upload_jobs:
            upload_jobs[self.job_id].stage = stage
            upload_jobs[self.job_id].progress = progress
    
    def create_smart_chunks(self, text: str, curr_page_no: int, no_of_pages: int, metadata: Dict) -> List[Dict]:
        """Create overlapping chunks with metadata"""
        words = text.split()
        chunks = []
        no_of_words = len(words)
        
        if len(words) <= settings.CHUNK_SIZE:
            return [{
                "content": text,
                "metadata": self._clean_metadata(metadata),
                "word_count": len(words)
            }]
        
        chunk_index = 0
        start = 0
        
        while start < len(words):
            progress_from_previous_pages = (curr_page_no - 1) / no_of_pages
            progress_on_current_page = (start / no_of_words) * (1 / no_of_pages)
            self.update_progress(ProcessingStage.CREATING_CHUNKS, (progress_from_previous_pages + progress_on_current_page) * 100)
            
            end = min(start + settings.CHUNK_SIZE, len(words))
            chunk_words = words[start:end]
            chunk_text = ' '.join(chunk_words)
            
            chunk_meta = metadata.copy()
            chunk_meta["chunk_index"] = chunk_index
            
            chunks.append({
                "content": chunk_text,
                "metadata": self._clean_metadata(chunk_meta),
                "word_count": len(chunk_words)
            })
            
            chunk_index += 1
            start = end - settings.CHUNK_OVERLAP
            
            if start >= len(words) - settings.CHUNK_OVERLAP:
                break
        
        return chunks
    
    def extract_with_llamaparse(self, pdf_path: str) -> Dict[int, str]:
        """Primary extraction using LlamaParse"""
        if not self.has_llamaparse:
            return {}
        
        try:
            print(f"🔍 LlamaParse processing for {Path(pdf_path).name}...")
            self.update_progress(ProcessingStage.LLAMA_LOADING, 0)
            documents = self.llama_parser.load_data(pdf_path)
            
            page_texts = {}
            no_of_pages = len(documents)
            for doc in documents:
                self.update_progress(ProcessingStage.EXTRACTING_CONTENT, (documents.index(doc) / no_of_pages) * 100)
                page_num_str = doc.metadata.get('page_label', doc.metadata.get('page', '1'))
                
                try:
                    page_num = int(float(page_num_str))
                except (ValueError, TypeError, AttributeError):
                    page_num = 1
                
                if page_num not in page_texts:
                    page_texts[page_num] = []
                
                page_texts[page_num].append(doc.text)
            
            result = {p: self.clean_text("\n\n".join(texts)) for p, texts in page_texts.items()}
            print(f"✅ LlamaParse extracted {len(result)} pages.")
            return result
            
        except Exception as e:
            print(f"❌ LlamaParse FATAL error on {Path(pdf_path).name}: {e}")
            traceback.print_exc()
            return {}
    
    def extract_with_pymupdf(self, pdf_path: str) -> Dict[int, str]:
        """Fallback extraction using PyMuPDF"""
        page_texts = {}
        try:
            self.update_progress(ProcessingStage.PYMUPDF_LOADING, 0)
            doc = fitz.open(pdf_path)
            no_of_pages = doc.page_count
            for page_num in range(doc.page_count):
                self.update_progress(ProcessingStage.EXTRACTING_CONTENT, (page_num / no_of_pages) * 100)
                page = doc[page_num]
                text = page.get_text()
                if text and text.strip():
                    page_texts[page_num + 1] = self.clean_text(text)
            doc.close()
            print(f"✅ PyMuPDF extracted {len(page_texts)} pages")
        except Exception as e:
            print(f"❌ PyMuPDF error: {e}")
        return page_texts
    
    def extract_with_tesseract(self, pdf_path: str) -> Dict[int, str]:
        """Final fallback OCR using Tesseract"""
        page_texts = {}
        try:
            print(f"🔎 Tesseract OCR processing...")
            self.update_progress(ProcessingStage.TESSERACT_LOADING, 0)
            doc = fitz.open(pdf_path)
            no_of_pages = doc.page_count
            for i, page in enumerate(doc):
                self.update_progress(ProcessingStage.EXTRACTING_CONTENT, (i / no_of_pages) * 100)
                pix = page.get_pixmap(dpi=200)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                text = pytesseract.image_to_string(img)
                if text and text.strip():
                    page_texts[i + 1] = self.clean_text(text)
            print(f"✅ Tesseract OCR extracted {len(page_texts)} pages.")
        except Exception as e:
            print(f"❌ Tesseract OCR error: {e}")
            traceback.print_exc()
        return page_texts
    
    def extract_tables(self, pdf_path: str) -> List[Dict]:
        """Extract tables using pdfplumber"""
        tables = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
                for page_num, page in enumerate(pdf.pages, start=1):
                    page_tables = page.extract_tables() or []
                    no_of_tables = len(page_tables)
                    for table_idx, table in enumerate(page_tables):
                        progress_from_previous_pages = (page_num - 1) / total_pages
                        progress_on_current_page = ((table_idx + 1) / no_of_tables) * (1 / total_pages)
                        total_progress = progress_from_previous_pages + progress_on_current_page
                        self.update_progress(ProcessingStage.EXTRACTING_TABLES, total_progress * 100)
                        if len(table) < 2:
                            continue
                        headers = table[0] if table[0] else []
                        table_text = f"Table {table_idx + 1} on page {page_num}:\n"
                        if headers:
                            table_text += "Headers: " + " | ".join(str(h) for h in headers if h) + "\n"
                        for row in table[1:]:
                            if not any(row):
                                continue
                            row_text = " | ".join(str(cell) if cell else "" for cell in row)
                            table_text += row_text + "\n"
                        tables.append({
                            "content": self.clean_text(table_text),
                            "page": page_num,
                            "type": "table",
                            "table_index": table_idx
                        })
            print(f"✅ Extracted {len(tables)} tables")
        except Exception as e:
            print(f"⚠️  Table extraction error: {e}")
        return tables
    
    def process_pdf(self, job_id: str, pdf_path: str, doc_id: str, filename: str) -> Tuple[List[Dict], Dict]:
        """Main PDF processing pipeline"""
        self.job_id = job_id
        print(f"\n{'='*60}\n📄 Processing PDF: {filename}\n{'='*60}")
        start_time = time.time()
        
        page_texts = self.extract_with_llamaparse(pdf_path)
        if not page_texts:
            print("⚠️  LlamaParse failed, attempting PyMuPDF fallback...")
            page_texts = self.extract_with_pymupdf(pdf_path)
        if not page_texts:
            print("⚠️  PyMuPDF also failed, attempting Tesseract OCR fallback...")
            page_texts = self.extract_with_tesseract(pdf_path)
        if not page_texts:
            raise Exception("Failed to extract any text from PDF")
        
        tables = self.extract_tables(pdf_path)
        
        all_chunks = []
        no_of_pages = len(page_texts)
        for page_num, text in page_texts.items():
            if not text.strip():
                continue
            base_metadata = {"doc_id": str(doc_id), "source": str(filename), "page": str(page_num), "type": "text", "doc_type": "pdf"}
            chunks = self.create_smart_chunks(text, page_num, no_of_pages, base_metadata)
            all_chunks.extend(chunks)
        
        for table in tables:
            table_meta = {"doc_id": str(doc_id), "source": str(filename), "page": str(table["page"]), "type": "table", "doc_type": "pdf", "table_index": str(table.get("table_index", 0))}
            all_chunks.append({"content": table["content"], "metadata": self._clean_metadata(table_meta), "word_count": len(table["content"].split())})
        
        if len(all_chunks) > settings.MAX_CHUNKS_PER_DOCUMENT:
            print(f"⚠️  Limiting to {settings.MAX_CHUNKS_PER_DOCUMENT} chunks")
            all_chunks = all_chunks[:settings.MAX_CHUNKS_PER_DOCUMENT]
        
        stats = {"total_chunks": len(all_chunks), "pages": len(page_texts), "tables": len(tables), "processing_time": time.time() - start_time}
        print(f"✅ Created {stats['total_chunks']} chunks from {stats['pages']} pages")
        print(f"⏱️  Processing time: {stats['processing_time']:.2f}s\n")
        
        return all_chunks, stats


# ============================================================================
# EXCEL PROCESSOR
# ============================================================================

class ExcelProcessor:
    """Enhanced Excel processing with multi-sheet and table extraction"""
    
    job_id = None
    
    def __init__(self, embedding_model, tokenizer):
        self.embedding_model = embedding_model
        self.tokenizer = tokenizer
        
        if not HAS_EXCEL_LIBS:
            self.has_excel_libs = False
            print("⚠️  Excel processing libraries not available")
        else:
            self.has_excel_libs = True
            print("✅ Excel processing libraries initialized (pandas, openpyxl)")
    
    def clean_text(self, text: str) -> str:
        """Clean and normalize text"""
        if not text:
            return ""
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r'[^\w\s\.\,\;\:\!\?\-\(\)\[\]\"\'\/\@\#\$\%\&\*\+\=]', ' ', text)
        return text.strip()
    
    def _clean_metadata(self, metadata: Dict) -> Dict:
        """Clean metadata for ChromaDB compatibility"""
        cleaned = {}
        for k, v in metadata.items():
            str_val = str(v)
            str_val = re.sub(r'[^\w\s\-\.\,\/]', '_', str_val)
            str_val = str_val.strip()
            cleaned[k] = str_val if str_val else "unknown"
        return cleaned
    
    def update_progress(self, stage: ProcessingStage, progress: float) -> None:
        print(f"📊 Progress: {stage} {progress:.1f}%")
        if not self.job_id:
            return
        if self.job_id in upload_jobs:
            upload_jobs[self.job_id].stage = stage
            upload_jobs[self.job_id].progress = progress
    
    def create_smart_chunks(self, text: str, curr_sheet_idx: int, no_of_sheets: int, metadata: Dict) -> List[Dict]:
        """Create overlapping chunks with metadata"""
        words = text.split()
        chunks = []
        no_of_words = len(words)
        
        if len(words) <= settings.CHUNK_SIZE:
            return [{
                "content": text,
                "metadata": self._clean_metadata(metadata),
                "word_count": len(words)
            }]
        
        chunk_index = 0
        start = 0
        
        while start < len(words):
            progress_from_previous_sheets = curr_sheet_idx / no_of_sheets if no_of_sheets > 0 else 0
            progress_on_current_sheet = (start / no_of_words) * (1 / no_of_sheets) if no_of_sheets > 0 else 0
            self.update_progress(ProcessingStage.CREATING_CHUNKS, (progress_from_previous_sheets + progress_on_current_sheet) * 100)
            end = min(start + settings.CHUNK_SIZE, len(words))
            chunk_words = words[start:end]
            chunk_text = ' '.join(chunk_words)
            
            chunk_meta = metadata.copy()
            chunk_meta["chunk_index"] = chunk_index
            
            chunks.append({
                "content": chunk_text,
                "metadata": self._clean_metadata(chunk_meta),
                "word_count": len(chunk_words)
            })
            
            chunk_index += 1
            start = end - settings.CHUNK_OVERLAP
            
            if start >= len(words) - settings.CHUNK_OVERLAP:
                break
        
        return chunks
    
    def extract_with_pandas(self, excel_path: str) -> Dict[str, str]:
        """Primary extraction using pandas for all sheets"""
        if not self.has_excel_libs:
            return {}
        
        try:
            print(f"🔍 Pandas processing for {Path(excel_path).name}...")
            self.update_progress(ProcessingStage.EXTRACTING_CONTENT, 0)
            
            # Check if it's an old .xls file and we need xlrd
            if excel_path.lower().endswith('.xls'):
                try:
                    # Try with xlrd engine for old Excel files
                    excel_file = pd.ExcelFile(excel_path, engine='xlrd')
                except ImportError:
                    print("❌ xlrd not available for .xls files. Install with: pip install xlrd")
                    return {}
                except Exception as e:
                    print(f"❌ xlrd failed to read .xls file: {e}")
                    return {}
            else:
                # Use default engine for .xlsx files
                excel_file = pd.ExcelFile(excel_path)
            
            sheet_texts = {}
            
            no_of_sheets = len(excel_file.sheet_names)
            print(f"📑 Found {no_of_sheets} sheets: {excel_file.sheet_names}")
            
            for idx, sheet_name in enumerate(excel_file.sheet_names):
                self.update_progress(ProcessingStage.EXTRACTING_CONTENT, (idx / no_of_sheets) * 100)
                
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    print(f"   Sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
                    
                    df = df.ffill(axis=0)
                    df = df.bfill().fillna('')
                    
                    text_parts = []
                    text_parts.append(f"Sheet: {sheet_name}\n")
                    
                    if not df.empty:
                        headers = " | ".join(str(col) for col in df.columns)
                        text_parts.append(f"Headers: {headers}\n")
                        
                        for row_idx, row in df.iterrows():
                            row_text = " | ".join(str(val) for val in row.values if str(val).strip())
                            if row_text.strip():
                                text_parts.append(row_text)
                    
                    sheet_text = "\n".join(text_parts)
                    if sheet_text.strip():
                        sheet_texts[sheet_name] = self.clean_text(sheet_text)
                
                except Exception as e:
                    print(f"⚠️  Error reading sheet '{sheet_name}': {e}")
                    continue
            
            print(f"✅ Pandas extracted {len(sheet_texts)} sheets.")
            return sheet_texts
            
        except ImportError as e:
            if "xlrd" in str(e):
                print(f"❌ Pandas extraction error: {e}")
                print("💡 For .xls files, install: pip install xlrd")
            else:
                print(f"❌ Pandas extraction error: {e}")
                traceback.print_exc()
            return {}
        except Exception as e:
            print(f"❌ Pandas extraction error: {e}")
            traceback.print_exc()
            return {}
    
    def extract_with_openpyxl(self, excel_path: str) -> Dict[str, str]:
        """Fallback extraction using openpyxl for direct cell access"""
        if not self.has_excel_libs:
            return {}
        
        try:
            from openpyxl import load_workbook
            from openpyxl.cell.cell import MergedCell
            
            print(f"🔍 OpenPyXL processing for {Path(excel_path).name}...")
            self.update_progress(ProcessingStage.EXTRACTING_CONTENT, 0)
            
            workbook = load_workbook(filename=excel_path, data_only=True)
            sheet_texts = {}
            
            no_of_sheets = len(workbook.sheetnames)
            print(f"📑 Found {no_of_sheets} sheets: {workbook.sheetnames}")
            
            for idx, sheet_name in enumerate(workbook.sheetnames):
                self.update_progress(ProcessingStage.EXTRACTING_CONTENT, (idx / no_of_sheets) * 100)
                
                try:
                    sheet = workbook[sheet_name]
                    text_parts = []
                    text_parts.append(f"Sheet: {sheet_name}\n")
                    
                    for row in sheet.iter_rows():
                        row_values = []
                        for cell in row:
                            if isinstance(cell, MergedCell):
                                continue
                            if cell.value is not None:
                                row_values.append(str(cell.value))
                        
                        if row_values:
                            row_text = " | ".join(row_values)
                            text_parts.append(row_text)
                    
                    sheet_text = "\n".join(text_parts)
                    if sheet_text.strip():
                        sheet_texts[sheet_name] = self.clean_text(sheet_text)
                
                except Exception as e:
                    print(f"⚠️  Error reading sheet '{sheet_name}': {e}")
                    continue
            
            workbook.close()
            print(f"✅ OpenPyXL extracted {len(sheet_texts)} sheets.")
            return sheet_texts
            
        except Exception as e:
            print(f"❌ OpenPyXL extraction error: {e}")
            traceback.print_exc()
            return {}
    
    def extract_tables(self, excel_path: str) -> List[Dict]:
        """Extract structured tables from Excel sheets"""
        tables = []
        
        if not self.has_excel_libs:
            return tables
        
        try:
            excel_file = pd.ExcelFile(excel_path)
            no_of_sheets = len(excel_file.sheet_names)
            
            for sheet_idx, sheet_name in enumerate(excel_file.sheet_names):
                progress = (sheet_idx / no_of_sheets) * 100
                self.update_progress(ProcessingStage.EXTRACTING_TABLES, progress)
                
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    if df.empty or len(df) < 2:
                        continue
                    
                    table_text = f"Table from sheet '{sheet_name}':\n"
                    headers = " | ".join(str(col) for col in df.columns)
                    table_text += f"Headers: {headers}\n"
                    
                    for _, row in df.iterrows():
                        row_text = " | ".join(str(val) if pd.notna(val) else "" for val in row.values)
                        if row_text.strip():
                            table_text += row_text + "\n"
                    
                    tables.append({
                        "content": self.clean_text(table_text),
                        "sheet": sheet_name,
                        "type": "table",
                        "rows": len(df),
                        "columns": len(df.columns)
                    })
                    
                except Exception as e:
                    print(f"⚠️  Error extracting table from sheet '{sheet_name}': {e}")
                    continue
            
            print(f"✅ Extracted {len(tables)} tables")
            
        except Exception as e:
            print(f"⚠️  Table extraction error: {e}")
            traceback.print_exc()
        
        return tables
    
    def process_excel(self, job_id: str, excel_path: str, doc_id: str, filename: str) -> Tuple[List[Dict], Dict]:
        """Main Excel processing pipeline"""
        self.job_id = job_id
        print(f"\n{'='*60}\n📊 Processing Excel: {filename}\n{'='*60}")
        start_time = time.time()
        
        if not self.has_excel_libs:
            raise Exception("Excel processing libraries not available. Install: pip install pandas openpyxl")
        
        # Check file extension and provide specific guidance
        if excel_path.lower().endswith('.xls'):
            print("📝 Processing old Excel format (.xls) - requires xlrd library")
        
        # Extract text from all sheets
        sheet_texts = self.extract_with_pandas(excel_path)
        if not sheet_texts:
            print("⚠️  Pandas extraction failed, attempting OpenPyXL fallback...")
            # Only try OpenPyXL for .xlsx files
            if not excel_path.lower().endswith('.xls'):
                sheet_texts = self.extract_with_openpyxl(excel_path)
            else:
                print("⚠️  OpenPyXL doesn't support .xls files. Please install xlrd or convert to .xlsx")
        
        if not sheet_texts:
            file_ext = Path(excel_path).suffix.lower()
            if file_ext == '.xls':
                raise Exception(f"Failed to extract data from {file_ext} file. Please install xlrd: pip install xlrd")
            else:
                raise Exception("Failed to extract any data from Excel file")
        
        tables = self.extract_tables(excel_path)
        
        all_chunks = []
        no_of_sheets = len(sheet_texts)
        
        for sheet_idx, (sheet_name, text) in enumerate(sheet_texts.items()):
            if not text.strip():
                continue
            
            base_metadata = {
                "doc_id": str(doc_id),
                "source": str(filename),
                "sheet": str(sheet_name),
                "type": "text",
                "doc_type": "excel"
            }
            
            chunks = self.create_smart_chunks(text, sheet_idx, no_of_sheets, base_metadata)
            all_chunks.extend(chunks)
        
        for table in tables:
            table_meta = {
                "doc_id": str(doc_id),
                "source": str(filename),
                "sheet": str(table["sheet"]),
                "type": "table",
                "doc_type": "excel",
                "rows": str(table.get("rows", 0)),
                "columns": str(table.get("columns", 0))
            }
            
            all_chunks.append({
                "content": table["content"],
                "metadata": self._clean_metadata(table_meta),
                "word_count": len(table["content"].split())
            })
        
        if len(all_chunks) > settings.MAX_CHUNKS_PER_DOCUMENT:
            print(f"⚠️  Limiting to {settings.MAX_CHUNKS_PER_DOCUMENT} chunks")
            all_chunks = all_chunks[:settings.MAX_CHUNKS_PER_DOCUMENT]
        
        stats = {
            "total_chunks": len(all_chunks),
            "sheets": len(sheet_texts),
            "tables": len(tables),
            "processing_time": time.time() - start_time
        }
        
        print(f"✅ Created {stats['total_chunks']} chunks from {stats['sheets']} sheets")
        print(f"⏱️  Processing time: {stats['processing_time']:.2f}s\n")
        
        return all_chunks, stats


# ============================================================================
# HTML PROCESSOR
# ============================================================================

class HTMLProcessor:
    """Enhanced HTML processing with table, link, and metadata extraction"""
    
    job_id = None
    
    def __init__(self, embedding_model, tokenizer):
        self.embedding_model = embedding_model
        self.tokenizer = tokenizer
        
        if not HAS_HTML_LIBS:
            self.has_html_libs = False
            print("⚠️  HTML processing libraries not available. Install: pip install beautifulsoup4")
        else:
            self.has_html_libs = True
            print("✅ HTML processing libraries initialized (BeautifulSoup4)")
    
    def clean_text(self, text: str) -> str:
        """Clean and normalize text"""
        if not text:
            return ""
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r'[^\w\s\.\,\;\:\!\?\-\(\)\[\]\"\'\/\@\#\$\%\&\*\+\=]', ' ', text)
        return text.strip()
    
    def _clean_metadata(self, metadata: Dict) -> Dict:
        """Clean metadata for ChromaDB compatibility"""
        cleaned = {}
        for k, v in metadata.items():
            str_val = str(v)
            str_val = re.sub(r'[^\w\s\-\.\,\/]', '_', str_val)
            str_val = str_val.strip()
            cleaned[k] = str_val if str_val else "unknown"
        return cleaned
    
    def update_progress(self, stage: ProcessingStage, progress: float) -> None:
        print(f"🌐 Progress: {stage} {progress:.1f}%")
        if not self.job_id:
            return
        if self.job_id in upload_jobs:
            upload_jobs[self.job_id].stage = stage
            upload_jobs[self.job_id].progress = progress
    
    def extract_title(self, soup) -> str:
        """Extract page title from HTML"""
        title = soup.find('title')
        if title:
            return title.get_text(strip=True)
        h1 = soup.find('h1')
        if h1:
            return h1.get_text(strip=True)
        return "Unknown"
    
    def extract_metadata(self, soup) -> Dict:
        """Extract metadata from HTML head"""
        metadata = {}
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        if meta_desc and meta_desc.get('content'):
            metadata['description'] = meta_desc.get('content')
        meta_keywords = soup.find('meta', attrs={'name': 'keywords'})
        if meta_keywords and meta_keywords.get('content'):
            metadata['keywords'] = meta_keywords.get('content')
        return metadata
    
    def extract_headings(self, soup) -> List[Dict]:
        """Extract heading structure from HTML"""
        headings = []
        for tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            for heading in soup.find_all(tag):
                text = heading.get_text(strip=True)
                if text:
                    headings.append({'level': int(tag[1]), 'text': text})
        return headings
    
    def extract_tables(self, soup) -> List[Dict]:
        """Extract tables from HTML"""
        tables = []
        for table_idx, table in enumerate(soup.find_all('table')):
            try:
                rows = table.find_all('tr')
                if len(rows) < 1:
                    continue
                
                table_data = []
                headers = []
                first_row = rows[0].find_all(['th', 'td'])
                for cell in first_row:
                    headers.append(cell.get_text(strip=True))
                
                if headers:
                    table_data.append(headers)
                
                start_idx = 1 if headers and rows[0].find('th') else 0
                for row in rows[start_idx:]:
                    cells = row.find_all(['td', 'th'])
                    if cells:
                        row_data = [cell.get_text(strip=True) for cell in cells]
                        table_data.append(row_data)
                
                if len(table_data) < 2:
                    continue
                
                table_text = f"Table {table_idx + 1}:\n"
                if headers:
                    table_text += "Headers: " + " | ".join(headers) + "\n"
                
                for row_data in table_data[1:]:
                    row_text = " | ".join(row_data)
                    if row_text.strip():
                        table_text += row_text + "\n"
                
                tables.append({
                    "content": self.clean_text(table_text),
                    "index": table_idx,
                    "rows": len(table_data) - 1,
                    "columns": len(headers) if headers else len(table_data[1]) if table_data else 0,
                    "type": "table"
                })
                
            except Exception as e:
                print(f"⚠️  Error extracting table {table_idx}: {e}")
                continue
        
        print(f"✅ Extracted {len(tables)} tables")
        return tables
    
    def extract_links(self, soup) -> List[Dict]:
        """Extract all links from HTML"""
        links = []
        for link in soup.find_all('a', href=True):
            url = link.get('href')
            text = link.get_text(strip=True)
            if url or text:
                links.append({'url': url or '', 'text': text or '[No text]', 'title': link.get('title', '')})
        print(f"✅ Extracted {len(links)} links")
        return links
    
    def extract_images(self, soup) -> List[Dict]:
        """Extract image metadata from HTML"""
        images = []
        for img in soup.find_all('img'):
            images.append({
                'src': img.get('src', ''),
                'alt': img.get('alt', ''),
                'title': img.get('title', '')
            })
        print(f"✅ Extracted {len(images)} images")
        return images
    
    def extract_main_content(self, soup) -> str:
        """Extract main content from HTML"""
        for script in soup(['script', 'style', 'noscript']):
            script.decompose()
        
        content = soup.find('main')
        if not content:
            article = soup.find('article')
            if article:
                content = article
            else:
                for div in soup.find_all('div', class_=re.compile('content|body|main|page|wrapper', re.I)):
                    content = div
                    break
        
        if not content:
            content = soup.find('body') or soup
        
        text = content.get_text(separator='\n', strip=True)
        return self.clean_text(text)
    
    def create_smart_chunks(self, text: str, curr_section_idx: int, no_of_sections: int, metadata: Dict) -> List[Dict]:
        """Create overlapping chunks with metadata"""
        words = text.split()
        chunks = []
        
        no_of_words = len(words)
        
        if len(words) <= settings.CHUNK_SIZE:
            return [{
                "content": text,
                "metadata": self._clean_metadata(metadata),
                "word_count": len(words)
            }]
        
        chunk_index = 0
        start = 0
        
        while start < len(words):
            progress_from_previous = curr_section_idx / no_of_sections if no_of_sections > 0 else 0
            progress_on_current = (start / no_of_words) * (1 / no_of_sections) if no_of_sections > 0 else 0
            self.update_progress(ProcessingStage.CREATING_CHUNKS, (progress_from_previous + progress_on_current) * 100)
            
            end = min(start + settings.CHUNK_SIZE, len(words))
            chunk_words = words[start:end]
            chunk_text = ' '.join(chunk_words)
            
            chunk_meta = metadata.copy()
            chunk_meta["chunk_index"] = chunk_index
            
            chunks.append({
                "content": chunk_text,
                "metadata": self._clean_metadata(chunk_meta),
                "word_count": len(chunk_words)
            })
            
            chunk_index += 1
            start = end - settings.CHUNK_OVERLAP
            
            if start >= len(words) - settings.CHUNK_OVERLAP:
                break
        
        return chunks
    
    def parse_html_file(self, html_path: str) -> Tuple:
        """Parse HTML file and return BeautifulSoup object"""
        try:
            with open(html_path, 'r', encoding='utf-8') as f:
                content = f.read()
        except UnicodeDecodeError:
            try:
                with open(html_path, 'r', encoding='latin-1') as f:
                    content = f.read()
            except:
                with open(html_path, 'r', encoding='iso-8859-1') as f:
                    content = f.read()
        
        soup = BeautifulSoup(content, 'html.parser')
        return soup, content
    
    def process_html(self, job_id: str, html_path: str, doc_id: str, filename: str) -> Tuple[List[Dict], Dict]:
        """Main HTML processing pipeline"""
        self.job_id = job_id
        print(f"\n{'='*60}\n🌐 Processing HTML: {filename}\n{'='*60}")
        start_time = time.time()
        
        if not self.has_html_libs:
            raise Exception("HTML processing libraries not available. Install: pip install beautifulsoup4")
        
        if not Path(html_path).exists():
            raise FileNotFoundError(f"HTML file not found: {html_path}")
        
        try:
            self.update_progress(ProcessingStage.HTML_LOADING, 0)
            soup, raw_content = self.parse_html_file(html_path)
            print(f"✅ HTML file loaded: {Path(html_path).name}")
        except Exception as e:
            print(f"❌ Error loading HTML file: {e}")
            raise
        
        self.update_progress(ProcessingStage.PARSING_CONTENT, 10)
        page_title = self.extract_title(soup)
        page_metadata = self.extract_metadata(soup)
        headings = self.extract_headings(soup)
        
        print(f"📄 Page Title: {page_title}")
        print(f"📋 Headings found: {len(headings)}")
        
        self.update_progress(ProcessingStage.PARSING_CONTENT, 30)
        main_content = self.extract_main_content(soup)
        print(f"✅ Main content extracted: {len(main_content.split())} words")
        
        self.update_progress(ProcessingStage.EXTRACTING_TABLES, 50)
        tables = self.extract_tables(soup)
        
        self.update_progress(ProcessingStage.EXTRACTING_LINKS, 70)
        links = self.extract_links(soup)
        images = self.extract_images(soup)
        
        all_chunks = []
        base_metadata = {
            "doc_id": str(doc_id),
            "source": str(filename),
            "page_title": str(page_title),
            "type": "text",
            "doc_type": "html"
        }
        
        chunks = self.create_smart_chunks(main_content, 0, 1, base_metadata)
        all_chunks.extend(chunks)
        
        for table in tables:
            table_meta = {
                "doc_id": str(doc_id),
                "source": str(filename),
                "page_title": str(page_title),
                "type": "table",
                "doc_type": "html",
                "table_index": str(table['index']),
                "rows": str(table['rows']),
                "columns": str(table['columns'])
            }
            
            all_chunks.append({
                "content": table["content"],
                "metadata": self._clean_metadata(table_meta),
                "word_count": len(table["content"].split())
            })
        
        if len(all_chunks) > settings.MAX_CHUNKS_PER_DOCUMENT:
            print(f"⚠️  Limiting to {settings.MAX_CHUNKS_PER_DOCUMENT} chunks")
            all_chunks = all_chunks[:settings.MAX_CHUNKS_PER_DOCUMENT]
        
        stats = {
            "total_chunks": len(all_chunks),
            "page_title": page_title,
            "tables": len(tables),
            "links": len(links),
            "images": len(images),
            "headings": len(headings),
            "main_content_words": len(main_content.split()),
            "processing_time": time.time() - start_time
        }
        
        print(f"✅ Created {stats['total_chunks']} chunks")
        print(f"⏱️  Processing time: {stats['processing_time']:.2f}s\n")
        
        return all_chunks, stats


# ============================================================================
# ARCHIVE PROCESSOR
# ============================================================================

class ArchiveProcessor:
    """
    Process archive files (ZIP, RAR, TAR, GZIP, 7Z) by extracting and
    recursively processing contained files.

    Archives are extracted to a temporary directory and their contents are
    processed by the appropriate file-type processor (PDF, Excel, HTML, etc.).

    Supports:
    - ZIP archives (.zip)
    - RAR archives (.rar)
    - TAR archives (.tar, .tar.gz, .tar.bz2)
    - 7-Zip archives (.7z)
    - Nested archives (recursively with depth limit)

    Features:
    - Automatic format detection
    - Safety limits (max recursion depth, max files, max size)
    - Graceful error handling (corrupted files don't stop processing)
    - Metadata tracking (preserves archive path in chunk metadata)
    """

    job_id = None
    _recursion_depth = 0
    _max_recursion_depth = 3

    def __init__(self, embedding_model, tokenizer, pdf_processor=None, excel_processor=None, html_processor=None, archive_processor=None):
        """
        Initialize the ArchiveProcessor.

        Args:
            embedding_model: Embedding model for chunking
            tokenizer: Tokenizer for text processing
            pdf_processor: PDFProcessor instance (created if not provided)
            excel_processor: ExcelProcessor instance (created if not provided)
            html_processor: HTMLProcessor instance (created if not provided)
            archive_processor: ArchiveProcessor instance for recursive processing
        """
        self.embedding_model = embedding_model
        self.tokenizer = tokenizer

        # Import here to avoid circular imports
        from app.modules.askai.services.archive_utils import (
            extract_archive,
            detect_archive_type,
            is_archive,
        )

        self.extract_archive = extract_archive
        self.detect_archive_type = detect_archive_type
        self.is_archive = is_archive

        # Initialize dependent processors (lazy initialization to avoid circular deps)
        self.pdf_processor = pdf_processor
        self.excel_processor = excel_processor
        self.html_processor = html_processor
        self.archive_processor = archive_processor

        if not HAS_PDF_LIBS:
            print("⚠️  PDF processing libraries not available for archive contents")

    def _get_processors(self):
        """
        Lazy initialization of dependent processors.
        This avoids circular imports when ArchiveProcessor is initialized.
        """
        if not self.pdf_processor:
            self.pdf_processor = PDFProcessor(self.embedding_model, self.tokenizer)
        if not self.excel_processor:
            self.excel_processor = ExcelProcessor(self.embedding_model, self.tokenizer)
        if not self.html_processor:
            self.html_processor = HTMLProcessor(self.embedding_model, self.tokenizer)
        if not self.archive_processor:
            # Recursive archive processing
            self.archive_processor = ArchiveProcessor(
                self.embedding_model,
                self.tokenizer,
                self.pdf_processor,
                self.excel_processor,
                self.html_processor,
            )

    def update_progress(self, stage: ProcessingStage, progress: float) -> None:
        print(f"📦 Progress: {stage} {progress:.1f}%")
        if not self.job_id:
            return
        if self.job_id in upload_jobs:
            upload_jobs[self.job_id].stage = stage
            upload_jobs[self.job_id].progress = progress

    def _clean_metadata(self, metadata: Dict) -> Dict:
        """Clean metadata for ChromaDB compatibility"""
        cleaned = {}
        for k, v in metadata.items():
            str_val = str(v)
            str_val = re.sub(r'[^\w\s\-\.\,\/]', '_', str_val)
            str_val = str_val.strip()
            cleaned[k] = str_val if str_val else "unknown"
        return cleaned

    def process_archive(self, job_id: str, archive_path: str, doc_id: str, filename: str) -> Tuple[List[Dict], Dict]:
        """
        Process an archive file by extracting and processing its contents.

        The archive is extracted to a temporary directory, and contained files
        are processed recursively based on their type. Supported formats:
        - PDFs → PDFProcessor
        - Excel files → ExcelProcessor
        - HTML files → HTMLProcessor
        - Archives → Recursive ArchiveProcessor (with depth limit)

        Metadata tracks the original archive filename and the path of each
        file within the archive structure.

        Args:
            job_id: Unique job identifier for progress tracking
            archive_path: Path to the archive file
            doc_id: Document ID for metadata
            filename: Original archive filename

        Returns:
            Tuple of (chunks, stats) where chunks are aggregated from all files

        Raises:
            Exception: If archive extraction fails or no files can be processed
        """
        self.job_id = job_id
        self._recursion_depth += 1

        print(f"\n{'='*60}\n📦 Processing Archive: {filename}\n{'='*60}")
        print(f"Recursion depth: {self._recursion_depth}/{self._max_recursion_depth}")

        start_time = time.time()

        # Safety check: prevent infinite recursion with nested archives
        if self._recursion_depth > self._max_recursion_depth:
            logger.warning(f"Archive recursion depth exceeded ({self._max_recursion_depth})")
            return [], {"error": "Max recursion depth exceeded", "processing_time": 0}

        try:
            # Create temporary directory for extraction
            archive_name = Path(filename).stem
            temp_extract_dir = Path(f"/tmp/archive_extract_{archive_name}_{uuid4()}")
            temp_extract_dir.mkdir(parents=True, exist_ok=True)

            self.update_progress(ProcessingStage.EXTRACTING_CONTENT, 0)

            # Extract archive
            logger.info(f"Extracting archive: {filename}")
            extracted_files = self.extract_archive(
                archive_path,
                str(temp_extract_dir),
                max_files=100,
                max_size_mb=500
            )

            if not extracted_files:
                logger.error(f"No files extracted from archive: {filename}")
                return [], {"error": "Archive extraction produced no files", "processing_time": 0}

            logger.info(f"Extracted {len(extracted_files)} files from archive")

            # Process extracted files
            all_chunks = []
            processed_count = 0
            failed_count = 0

            for file_idx, file_path in enumerate(extracted_files):
                progress = (file_idx / len(extracted_files)) * 100
                self.update_progress(ProcessingStage.EXTRACTING_CONTENT, progress)

                try:
                    file_ext = file_path.suffix.lower()

                    # Get file path relative to archive for metadata
                    relative_path = file_path.relative_to(temp_extract_dir)

                    logger.info(f"Processing extracted file: {relative_path}")

                    self._get_processors()

                    if file_ext == '.pdf':
                        chunks, _ = self.pdf_processor.process_pdf(
                            job_id, str(file_path), doc_id, file_path.name
                        )
                    elif file_ext in ['.xls', '.xlsx']:
                        chunks, _ = self.excel_processor.process_excel(
                            job_id, str(file_path), doc_id, file_path.name
                        )
                    elif file_ext == '.html':
                        chunks, _ = self.html_processor.process_html(
                            job_id, str(file_path), doc_id, file_path.name
                        )
                    elif self.is_archive(file_path):
                        # Recursively process nested archives
                        chunks, _ = self.archive_processor.process_archive(
                            job_id, str(file_path), doc_id, file_path.name
                        )
                    else:
                        logger.warning(f"Skipped unsupported file type: {file_path.name}")
                        failed_count += 1
                        continue

                    # Add archive metadata to chunks
                    for chunk in chunks:
                        chunk['metadata']['archive_filename'] = filename
                        chunk['metadata']['archive_path'] = str(relative_path)
                        chunk['metadata']['extraction_depth'] = self._recursion_depth

                    all_chunks.extend(chunks)
                    processed_count += 1

                except Exception as e:
                    logger.error(f"Failed to process {file_path.name} from archive: {e}")
                    failed_count += 1
                    # Continue with next file - don't fail entire archive

            # Cleanup extracted files
            try:
                shutil.rmtree(temp_extract_dir, ignore_errors=True)
            except Exception as e:
                logger.warning(f"Failed to clean up temp directory: {e}")

            stats = {
                "total_chunks": len(all_chunks),
                "files_processed": processed_count,
                "files_failed": failed_count,
                "archive_filename": filename,
                "recursion_depth": self._recursion_depth,
                "processing_time": time.time() - start_time
            }

            print(f"✅ Processed {processed_count}/{len(extracted_files)} files from archive")
            print(f"⏱️  Processing time: {stats['processing_time']:.2f}s\n")

            self._recursion_depth -= 1
            return all_chunks, stats

        except Exception as e:
            logger.error(f"Failed to process archive {filename}: {e}")
            traceback.print_exc()
            self._recursion_depth -= 1
            raise


# ============================================================================
# UNIFIED DOCUMENT SERVICE
# ============================================================================

class DocumentService:
    """
    Unified service to handle PDF, Excel, HTML, and Archive document processing.

    Routes files to appropriate processors based on file type:
    - .pdf → PDFProcessor
    - .xls, .xlsx → ExcelProcessor
    - .html → HTMLProcessor
    - .zip, .rar, .tar, .gz, .7z → ArchiveProcessor
    """

    def __init__(self, embedding_model=None, tokenizer=None):
        """Initialize all processors"""
        self.pdf_processor = PDFProcessor(embedding_model, tokenizer)
        self.excel_processor = ExcelProcessor(embedding_model, tokenizer)
        self.html_processor = HTMLProcessor(embedding_model, tokenizer)
        # Archive processor references other processors (initialized later to avoid circular deps)
        self.archive_processor = ArchiveProcessor(
            embedding_model,
            tokenizer,
            self.pdf_processor,
            self.excel_processor,
            self.html_processor,
        )
    
    def process_document(self, job_id: str, file_path: str, doc_id: str, filename: str, save_json: bool = True) -> Tuple[List[Dict], Dict]:
        """
        Process any document type and return chunks with stats.

        Supports multiple file types with automatic routing:
        - PDF (.pdf) → PDFProcessor (with LlamaParse OCR)
        - Excel (.xls, .xlsx) → ExcelProcessor (multi-sheet support)
        - HTML (.html, .htm) → HTMLProcessor (structure preservation)
        - Archives (.zip, .rar, .tar, .tar.gz, .7z) → ArchiveProcessor (recursive extraction)

        Archives are transparently extracted and their contents are processed
        as if they were individual files. Metadata preserves information about
        the archive and file location within the archive.

        Args:
            job_id: Unique job identifier for progress tracking
            file_path: Path to the file to process
            doc_id: Document ID for metadata
            filename: Original filename
            save_json: Whether to save chunks as JSON

        Returns:
            Tuple of (chunks, stats)

        Raises:
            ValueError: If file type is not supported
        """
        file_ext = Path(filename).suffix.lower()
        name_lower = filename.lower()

        print(f"\n🔄 Routing to appropriate processor based on file type: {file_ext}")

        try:
            if file_ext == '.pdf':
                chunks, stats = self.pdf_processor.process_pdf(job_id, file_path, doc_id, filename)
            elif file_ext in ['.xls', '.xlsx']:
                chunks, stats = self.excel_processor.process_excel(job_id, file_path, doc_id, filename)
            elif file_ext == '.html':
                chunks, stats = self.html_processor.process_html(job_id, file_path, doc_id, filename)
            elif (
                file_ext in ['.zip', '.rar', '.7z', '.gz', '.bz2']
                or name_lower.endswith('.tar.gz')
                or name_lower.endswith('.tar.bz2')
                or file_ext == '.tar'
                or name_lower.endswith('.tgz')
            ):
                # Archive formats: ZIP, RAR, 7Z, TAR variants, GZIP, BZIP2
                chunks, stats = self.archive_processor.process_archive(job_id, file_path, doc_id, filename)
            else:
                supported = ".pdf, .xls, .xlsx, .html, .zip, .rar, .tar, .tar.gz, .tar.bz2, .tgz, .7z"
                raise ValueError(f"Unsupported file type: {file_ext}. Supported: {supported}")

            # Optionally save chunks to JSON
            if save_json:
                json_filename = f"{Path(filename).stem}_chunks_output.json"
                with open(json_filename, "w", encoding="utf-8") as f:
                    json.dump(chunks, f, indent=2, ensure_ascii=False)
                print(f"💾 Saved chunks to: {json_filename}")

            return chunks, stats

        except Exception as e:
            print(f"❌ Document processing failed: {e}")
            traceback.print_exc()
            raise
